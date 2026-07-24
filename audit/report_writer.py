"""
==========================================================
Hospital Inventory Analytics System (HIAS)
Module      : Report Writer
Version     : 3.0.0
Author      : Mangesh Ukey
==========================================================
"""

from pathlib import Path
from datetime import datetime

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from audit.audit_utils import (
    OUTPUT_FOLDER,
    MAX_ROWS_PER_SHEET,
    log_info,
    log_error,
    PerformanceTimer,
)


class ReportWriter:
    """
    Enterprise Excel Report Writer
    """

    def __init__(self, report_name: str):

        self.report_name = report_name

        self.workbook = Workbook()

        self.current_sheet = self.workbook.active

        self.current_sheet.title = "Summary"

        self.current_row = 1

        self.output_file = (
            OUTPUT_FOLDER /
            f"{report_name}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        log_info(f"Report Writer Initialized : {report_name}")

    # ----------------------------------------------------

    def create_sheet(self, sheet_name: str):

        self.current_sheet = self.workbook.create_sheet(title=sheet_name)

        self.current_row = 1

        log_info(f"Sheet Created : {sheet_name}")

        return self.current_sheet
    
    # ----------------------------------------------------

    def write_dataframe(
        self,
        dataframe: pd.DataFrame,
        sheet_name: str,
        auto_format: bool = True,
    ):
        """
        Write DataFrame to Excel worksheet.

        Parameters
        ----------
        dataframe : pandas.DataFrame
        sheet_name : Sheet Name
        auto_format : Apply Formatting
        """

        timer = PerformanceTimer(
            f"Write Sheet : {sheet_name}"
        )

        timer.start()

        try:

            if dataframe is None:

                log_error(
                    f"{sheet_name} : DataFrame is None."
                )

                return

            if dataframe.empty:

                log_info(
                    f"{sheet_name} : No Records Found."
                )

                return

            sheet = self.create_sheet(sheet_name)

            # --------------------------
            # Header
            # --------------------------

            for col_no, column in enumerate(
                dataframe.columns,
                start=1,
            ):

                cell = sheet.cell(
                    row=1,
                    column=col_no,
                )

                cell.value = column

                cell.font = Font(
                    bold=True,
                    color="FFFFFF",
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="1F4E78",
                )

                cell.alignment = Alignment(
                    horizontal="center"
                )

            current_row = 2

            # --------------------------
            # Data
            # --------------------------

            for row in dataframe.itertuples(
                index=False,
                name=None,
            ):

                if current_row > MAX_ROWS_PER_SHEET:

                    sheet = self.create_sheet(
                        f"{sheet_name}_2"
                    )

                    current_row = 1

                    for col_no, column in enumerate(
                        dataframe.columns,
                        start=1,
                    ):

                        cell = sheet.cell(
                            row=1,
                            column=col_no,
                        )

                        cell.value = column

                        cell.font = Font(
                            bold=True,
                            color="FFFFFF",
                        )

                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color="1F4E78",
                        )

                    current_row = 2

                sheet.append(row)

                current_row += 1

            # --------------------------
            # Formatting
            # --------------------------

            if auto_format:

                sheet.freeze_panes = "A2"

                sheet.auto_filter.ref = (
                    sheet.dimensions
                )

                for column_cells in sheet.columns:

                    length = max(
                        len(str(cell.value))
                        if cell.value is not None
                        else 0
                        for cell in column_cells
                    )

                    letter = get_column_letter(
                        column_cells[0].column
                    )

                    sheet.column_dimensions[
                        letter
                    ].width = min(
                        max(length + 2, 12),
                        50,
                    )

            log_info(
                f"{sheet_name} Exported Successfully."
            )

            timer.stop()

        except Exception as ex:

            log_error(str(ex))

            raise
    # ----------------------------------------------------

    def save(self):

        timer = PerformanceTimer("Workbook Save")

        timer.start()

        try:

            self.workbook.save(self.output_file)

            log_info(f"Workbook Saved : {self.output_file}")

            timer.stop()

            return str(self.output_file)

        except Exception as ex:

            log_error(str(ex))

            raise
if __name__ == "__main__":

    df = pd.DataFrame({

        "Item": ["PCM", "Dolo"],

        "Qty": [100, 50],

        "Amount": [1200, 800]

    })

    writer = ReportWriter("Sample_Report")

    writer.write_dataframe(

        dataframe=df,

        sheet_name="Inventory"

    )

    file = writer.save()
    print(file)
    