"""
------------------------------------------------------------
Hospital Inventory Analytics System (HIAS)
Dashboard Module
------------------------------------------------------------
Author : Mangesh Ukey
Version : 1.0
------------------------------------------------------------
"""

import os
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment


class Dashboard:

    def __init__(self, workbook, output_folder):

        self.workbook = workbook
        self.output_folder = output_folder

        self.chart_folder = os.path.join(output_folder, "Charts")

        os.makedirs(self.chart_folder, exist_ok=True)

    # ---------------------------------------------------------
    # Dashboard Sheet
    # ---------------------------------------------------------
    def create_dashboard(
            self,
            daily_df,
            weekly_df,
            monthly_df,
            summary
    ):

        if "Dashboard" in self.workbook.sheetnames:
            del self.workbook["Dashboard"]

        ws = self.workbook.create_sheet("Dashboard", 0)

        # ---------------------------------------------------------
        # Title
        # ---------------------------------------------------------
        ws.merge_cells("A1:H1")

        ws["A1"] = "Hospital Inventory Analytics Dashboard"

        ws["A1"].font = Font(
            bold=True,
            size=20,
            color="FFFFFF"
        )

        ws["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        ws["A1"].alignment = Alignment(horizontal="center")

        # ---------------------------------------------------------
        # KPI Section
        # ---------------------------------------------------------

        row = 3

        for key, value in summary.items():

            ws.cell(row=row, column=1).value = key

            ws.cell(row=row, column=2).value = value

            ws.cell(row=row, column=1).font = Font(bold=True)

            row += 1

        # ---------------------------------------------------------
        # Charts
        # ---------------------------------------------------------

        self.purchase_chart(monthly_df)

        self.sales_chart(monthly_df)

        self.stock_value_chart(daily_df)

        self.top_items_chart(monthly_df)

        charts = [

            ("purchase.png", "A15"),

            ("sales.png", "J15"),

            ("stock_value.png", "A35"),

            ("top_items.png", "J35")

        ]

        for chart, cell in charts:

            path = os.path.join(
                self.chart_folder,
                chart
            )

            if os.path.exists(path):

                img = Image(path)

                img.width = 500

                img.height = 300

                ws.add_image(img, cell)

    # ---------------------------------------------------------
    # Purchase Trend
    # ---------------------------------------------------------
    def purchase_chart(self, monthly_df):

        data = (
            monthly_df
            .groupby("Month")["Purchase Qty"]
            .sum()
        )

        plt.figure(figsize=(7, 4))

        plt.plot(
            data.index.astype(str),
            data.values,
            marker="o"
        )

        plt.title("Monthly Purchase")

        plt.xlabel("Month")

        plt.ylabel("Quantity")

        plt.xticks(rotation=45)

        plt.tight_layout()

        plt.savefig(
            os.path.join(
                self.chart_folder,
                "purchase.png"
            )
        )

        plt.close()

    # ---------------------------------------------------------
    # Sales Trend
    # ---------------------------------------------------------
    def sales_chart(self, monthly_df):

        data = (
            monthly_df
            .groupby("Month")["Sales Qty"]
            .sum()
        )

        plt.figure(figsize=(7, 4))

        plt.plot(
            data.index.astype(str),
            data.values,
            marker="o"
        )

        plt.title("Monthly Sales")

        plt.xlabel("Month")

        plt.ylabel("Quantity")

        plt.xticks(rotation=45)

        plt.tight_layout()

        plt.savefig(
            os.path.join(
                self.chart_folder,
                "sales.png"
            )
        )

        plt.close()

    # ---------------------------------------------------------
    # Stock Value
    # ---------------------------------------------------------
    def stock_value_chart(self, daily_df):

        data = (
            daily_df
            .groupby("Date")["MRP Value"]
            .sum()
        )

        plt.figure(figsize=(7, 4))

        plt.plot(
            data.index.astype(str),
            data.values,
            marker="."
        )

        plt.title("Daily Stock Value")

        plt.xlabel("Date")

        plt.ylabel("MRP Value")

        plt.xticks(rotation=45)

        plt.tight_layout()

        plt.savefig(
            os.path.join(
                self.chart_folder,
                "stock_value.png"
            )
        )

        plt.close()

    # ---------------------------------------------------------
    # Top Selling Items
    # ---------------------------------------------------------
    def top_items_chart(self, monthly_df):

        data = (
            monthly_df
            .groupby("Item Name")["Sales Qty"]
            .sum()
            .sort_values(ascending=False)
            .head(10)
        )

        plt.figure(figsize=(7, 5))

        plt.barh(
            data.index,
            data.values
        )

        plt.title("Top 10 Selling Items")

        plt.tight_layout()

        plt.savefig(
            os.path.join(
                self.chart_folder,
                "top_items.png"
            )
        )

        plt.close()