"""
------------------------------------------------------------
Hospital Inventory Analytics System (HIAS)
Report Generator
------------------------------------------------------------
Author : Mangesh Ukey
Version : 1.0
------------------------------------------------------------
"""

import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter


class ReportGenerator:

    def __init__(self, output_folder):

        self.output_folder = output_folder

        self.wb = Workbook()

        # Remove default sheet
        self.wb.remove(self.wb.active)

        # Styles
        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        self.header_font = Font(
            color="FFFFFF",
            bold=True
        )

        self.data_font = Font(
            color="000000"
        )

        self.border = Border(

            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")

        )

    # --------------------------------------------------------
    # Add DataFrame as Worksheet
    # --------------------------------------------------------
    def add_sheet(self, sheet_name, dataframe):

        ws = self.wb.create_sheet(sheet_name)

        # Report Title
        ws["A1"] = "Hospital Inventory Analytics System"

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        ws["A2"] = f"Generated : {datetime.now()}"

        # Header
        row = 4

        for col, column_name in enumerate(dataframe.columns, start=1):

            cell = ws.cell(row=row, column=col)

            cell.value = column_name

            cell.fill = self.header_fill

            cell.font = self.header_font

            cell.alignment = Alignment(horizontal="center")

            cell.border = self.border

        # Data
        for r, values in enumerate(dataframe.values, start=row + 1):

            for c, value in enumerate(values, start=1):

                cell = ws.cell(row=r, column=c)

                cell.value = value

                cell.font = self.data_font

                cell.border = self.border

        # Freeze Header
        ws.freeze_panes = "A5"

        # Filter
        ws.auto_filter.ref = ws.dimensions

        # Auto Width
        self.auto_fit(ws)

    # --------------------------------------------------------
    # Auto Fit Columns
    # --------------------------------------------------------
    def auto_fit(self, ws):

        for column in ws.columns:

            max_length = 0

            letter = get_column_letter(column[0].column)

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except Exception:
                    pass

            ws.column_dimensions[letter].width = min(max_length + 3, 40)

    # --------------------------------------------------------
    # Summary Sheet
    # --------------------------------------------------------
    def add_summary(self, summary_dict):

        ws = self.wb.create_sheet("Summary")

        ws["A1"] = "Inventory Summary"

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        row = 3

        for key, value in summary_dict.items():

            ws.cell(row=row, column=1).value = key

            ws.cell(row=row, column=2).value = value

            ws.cell(row=row, column=1).font = Font(bold=True)

            row += 1

        self.auto_fit(ws)

    # --------------------------------------------------------
    # Save Workbook
    # --------------------------------------------------------
    def save(self):

        filename = os.path.join(

            self.output_folder,

            f"Hospital_Inventory_Report_"
            f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"

        )

        self.wb.save(filename)

        return filename

       # --------------------------------------------------------
      # Generate All Reports
      # --------------------------------------------------------
    def generate(
        self,
        daily,
        weekly,
        monthly,
        ledger,
        negative,
        zero,
        slow,
        summary
    ):

        print("Creating Daily Stock...")
        self.add_sheet("Daily Stock", daily)

        print("Creating Weekly Stock...")
        self.add_sheet("Weekly Stock", weekly)

        print("Creating Monthly Stock...")
        self.add_sheet("Monthly Stock", monthly)

        print("Creating Item Ledger...")
        self.add_sheet("Item Ledger", ledger)

        print("Creating Negative Stock...")
        self.add_sheet("Negative Stock", negative)

        print("Creating Zero Stock...")
        self.add_sheet("Zero Stock", zero)

        print("Creating Slow Moving...")
        self.add_sheet("Slow Moving", slow)

        print("Creating Summary...")
        self.add_summary(summary)

        print("Saving Excel...")

        filename = self.save()

        print(f"Excel Saved Successfully: {filename}")

        return filename